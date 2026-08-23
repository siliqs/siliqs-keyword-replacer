# -*- coding: utf-8 -*-
"""舊版 .doc / .xls 測試。

轉檔很貴（每次 5–20 秒），所以用 session fixture 讓每種格式只跑一次 round-trip，
所有斷言共用同一份產出。這也讓「轉檔壞掉」時只浪費一次逾時，不是每個測試各等一次。

沒裝 LibreOffice 就整批跳過——跳過不等於通過。
"""
import io
import os

import openpyxl
import pytest
from docx import Document

from conftest import SAMPLES, call
from src import legacy_office
from src.config import MatchOptions
from src.handlers import doc_file, xls_file

pytestmark = pytest.mark.skipif(not legacy_office.is_available(), reason="未安裝 LibreOffice")

DOC = os.path.join(SAMPLES, "split_runs.doc")
XLS = os.path.join(SAMPLES, "book.xls")


@pytest.fixture(scope="session")
def workdir(tmp_path_factory):
    return tmp_path_factory.mktemp("legacy")


@pytest.fixture(scope="session")
def doc_result(workdir):
    out_dir = workdir / "doc_out"
    out_dir.mkdir()
    dst = str(out_dir / "out.doc")
    count, message = call(doc_file, DOC, dst, ["機密"], MatchOptions())
    return {"dst": dst, "dir": str(out_dir), "count": count, "message": message}


@pytest.fixture(scope="session")
def doc_as_docx(doc_result, workdir):
    return legacy_office.convert(doc_result["dst"], "docx", str(workdir / "doc_check"))


@pytest.fixture(scope="session")
def xls_result(workdir):
    dst = str(workdir / "out.xls")
    count, message = call(xls_file, XLS, dst, ["機密"], MatchOptions())
    return {"dst": dst, "count": count, "message": message}


@pytest.fixture(scope="session")
def xls_as_xlsx(xls_result, workdir):
    return legacy_office.convert(xls_result["dst"], "xlsx", str(workdir / "xls_check"))


def test_doc_stays_doc_and_replaces(doc_result, doc_as_docx):
    assert os.path.splitext(doc_result["dst"])[1] == ".doc"        # R2：進出同格式
    assert doc_result["count"] >= 1
    assert "LibreOffice" in doc_result["message"]                   # 轉檔失真風險要講明

    text = "\n".join(p.text for p in Document(doc_as_docx).paragraphs)
    assert "機密" not in text
    assert "snoopy" in text


def test_doc_output_is_real_legacy_format(doc_result):
    """輸出必須是 OLE2 複合檔（舊版 .doc），不能是偷換成 docx 的 zip。"""
    header = io.open(doc_result["dst"], "rb").read(8)
    assert header == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"            # OLE2 magic
    assert not header.startswith(b"PK")


def test_no_intermediate_files_left_in_output(doc_result):
    """中間格式（docx/xlsx）絕不能落到輸出目錄。"""
    assert sorted(os.listdir(doc_result["dir"])) == ["out.doc"]


def test_source_untouched(doc_result):
    with io.open(DOC, "rb") as fh:
        assert fh.read(8) == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"    # R1：來源仍是原本的 .doc
    assert os.path.getsize(DOC) > 0


def test_xls_stays_xls_and_replaces(xls_result, xls_as_xlsx):
    assert xls_result["count"] >= 1
    workbook = openpyxl.load_workbook(xls_as_xlsx)
    values = [cell.value for sheet in workbook.worksheets
              for row in sheet.iter_rows() for cell in row if isinstance(cell.value, str)]
    assert not any("機密" in v for v in values)
    assert any("snoopy" in v for v in values)


def test_xls_keeps_sheets(xls_as_xlsx):
    assert openpyxl.load_workbook(xls_as_xlsx).sheetnames == ["資料", "第二頁"]


def test_skips_cleanly_without_libreoffice(tmp_path, monkeypatch):
    monkeypatch.setattr(legacy_office, "find_soffice", lambda: None)
    with pytest.raises(legacy_office.SofficeUnavailableError):
        doc_file.process(DOC, str(tmp_path / "out.doc"), ["機密"], MatchOptions())
