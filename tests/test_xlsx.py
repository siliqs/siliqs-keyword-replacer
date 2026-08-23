# -*- coding: utf-8 -*-
import io
import os

import openpyxl
import pytest

from conftest import SAMPLES, call
from src.config import MatchOptions
from src.handlers import xlsx_file

SRC = os.path.join(SAMPLES, "book.xlsx")


def _process(tmp_path, keywords=("機密",)):
    dst = str(tmp_path / "out.xlsx")
    count, _ = call(xlsx_file, SRC, dst, keywords, MatchOptions())
    return openpyxl.load_workbook(dst), count


def test_replaces_string_cells_across_sheets(tmp_path):
    wb, count = _process(tmp_path)
    assert wb["資料"]["A1"].value == "snoopy資料"
    assert wb["資料"]["A2"].value == "一般資料"
    assert wb["資料"]["A4"].value == "合併儲存格內的snoopy"
    assert wb["第二頁"]["A1"].value == "snoopy"
    assert count == 4                                    # A1 + A4 + 第二頁A1 + B1 公式內字串


def test_formula_string_literal_replaced_but_references_intact(tmp_path):
    wb, _ = _process(tmp_path)
    ws = wb["資料"]
    assert ws["B1"].value == '=CONCATENATE("snoopy-",A1)'   # 只換引號內的字串常數
    assert ws["B2"].value == "=SUM(C1:C9)"                  # 純參照公式不得動
    assert ws["C1"].value == 123                            # 數值不得動


def test_styles_widths_and_merges_preserved(tmp_path):
    wb, _ = _process(tmp_path)
    ws = wb["資料"]
    assert ws["A1"].font.bold is True
    assert ws["A1"].font.size == 14
    assert ws.column_dimensions["A"].width == 28
    assert "A4:B4" in [str(r) for r in ws.merged_cells.ranges]


def test_source_untouched(tmp_path):
    before = io.open(SRC, "rb").read()
    _process(tmp_path)
    assert io.open(SRC, "rb").read() == before              # R1


def test_workbook_with_chart_is_refused(tmp_path):
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference

    src = str(tmp_path / "chart.xlsx")
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "機密"
    ws["A2"] = 1
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=1, min_row=2, max_row=2))
    ws.add_chart(chart, "C1")
    wb.save(src)

    # 寧可 SKIP 也不默默把圖表弄丟（R5）
    with pytest.raises(xlsx_file.UnsupportedFeatureError):
        xlsx_file.process(src, str(tmp_path / "out.xlsx"), ["機密"], MatchOptions())


def test_image_warning_when_ocr_unavailable(tmp_path, monkeypatch):
    """有圖但 OCR 不可用：照樣處理文字，並在報表訊息留下警告，不得默默略過。"""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage

    from src import ocr

    png = str(tmp_path / "pic.png")
    PILImage.new("RGB", (60, 30), (240, 240, 240)).save(png)
    src = str(tmp_path / "with_image.xlsx")
    wb = Workbook()
    wb.active["A1"] = "機密"
    wb.active.add_image(XLImage(png), "C3")
    wb.save(src)

    monkeypatch.setattr(ocr, "is_available", lambda: False)
    count, message = xlsx_file.process(src, str(tmp_path / "out.xlsx"), ["機密"], MatchOptions())
    assert count == 1
    assert "Tesseract" in message
